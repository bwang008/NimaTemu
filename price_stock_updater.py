import pandas as pd
import shutil
from openpyxl import load_workbook
import warnings

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

class PriceStockUpdater:
    """Handles price and stock updates from PRICES.XLS file"""

    def __init__(self):
        self.prices_file = 'data/price/PRICES.XLS'
        self.price_template = 'data/temu_price_template.xlsx'
        self.stock_template = 'data/temu_stock_template.xlsx'
        self.price_output = 'output/temu_price_update.xlsx'
        self.stock_output = 'output/temu_stock_update.xlsx'

        # Load PRICES.XLS data
        self.prices_df = None
        self.load_prices_data()

    def load_prices_data(self):
        """Load and prepare PRICES.XLS data"""
        try:
            print("Loading PRICES.XLS data...")
            # Read starting from row 6 (header=5 means row 6 is the header)
            self.prices_df = pd.read_excel(self.prices_file, sheet_name='Sheet1', header=5)

            # Clean up the data
            self.prices_df = self.prices_df.dropna(subset=['Item #'])  # Remove rows without SKU

            # Convert relevant columns to appropriate types
            self.prices_df['On-hand Qty'] = pd.to_numeric(self.prices_df['On-hand Qty'], errors='coerce').fillna(0)
            self.prices_df['Sale Price'] = pd.to_numeric(self.prices_df['Sale Price'], errors='coerce')

            print(f"  Loaded {len(self.prices_df)} price records")
            print(f"  Columns: {list(self.prices_df.columns)}")

        except Exception as e:
            print(f"Error loading PRICES.XLS: {e}")
            self.prices_df = pd.DataFrame()

    def get_price_for_sku(self, sku):
        """Get sale price for a given SKU"""
        if self.prices_df is None or self.prices_df.empty:
            return None

        # Find matching SKU
        match = self.prices_df[self.prices_df['Item #'] == sku]
        if not match.empty:
            return match.iloc[0]['Sale Price']
        return None

    def get_stock_for_sku(self, sku):
        """Get stock quantity for a given SKU"""
        if self.prices_df is None or self.prices_df.empty:
            return 0

        # Find matching SKU
        match = self.prices_df[self.prices_df['Item #'] == sku]
        if not match.empty:
            return match.iloc[0]['On-hand Qty']
        return 0

    def create_price_update_file(self, product_data, base_prices):
        """Create price update file from product data with chunking"""
        try:
            print("Creating price update files with chunking...")

            # Split data into chunks of 1000 records
            chunk_size = 1000
            total_records = len(product_data)
            chunks = []
            
            for i in range(0, total_records, chunk_size):
                chunk_end = min(i + chunk_size, total_records)
                chunk_data = product_data.iloc[i:chunk_end]
                chunk_prices = base_prices[i:chunk_end]
                chunks.append((chunk_data, chunk_prices))

            print(f"  Split {total_records} records into {len(chunks)} chunks")

            # Process each chunk
            for chunk_idx, (chunk_data, chunk_prices) in enumerate(chunks, 1):
                chunk_filename = self.price_output.replace('.xlsx', f'_{chunk_idx}.xlsx')
                print(f"  Creating chunk {chunk_idx}/{len(chunks)}: {len(chunk_data)} records -> {chunk_filename}")

                # Copy template
                shutil.copy2(self.price_template, chunk_filename)

                # Load workbook
                workbook = load_workbook(chunk_filename)
                sheet = workbook.active

                # Prepare data for this chunk
                price_data = []
                for i, (sku, base_price) in enumerate(zip(chunk_data['SKU'], chunk_prices)):
                    if pd.notna(sku) and str(sku).strip() != '':
                        # Get new price from PRICES.XLS
                        new_price = self.get_price_for_sku(str(sku))

                        # If no new price found, use base price
                        if new_price is None or pd.isna(new_price):
                            new_price = base_price

                        price_data.append({
                            'SKU ID': str(sku),
                            'Current base price': base_price,
                            'New base price': new_price
                        })

                # Write data starting from row 2 (row 1 has headers)
                for i, row_data in enumerate(price_data, 2):
                    sheet.cell(row=i, column=1, value=row_data['SKU ID'])
                    sheet.cell(row=i, column=2, value=row_data['Current base price'])
                    sheet.cell(row=i, column=3, value=row_data['New base price'])

                # Save workbook
                workbook.save(chunk_filename)
                workbook.close()

                print(f"    Created chunk {chunk_idx} with {len(price_data)} records")

            print(f"  Created {len(chunks)} price update files")
            return True

        except Exception as e:
            print(f"Error creating price update files: {e}")
            return False

    def create_stock_update_file(self, product_data):
        """Create stock update file from product data with chunking"""
        try:
            print("Creating stock update files with chunking...")

            # Split data into chunks of 1000 records
            chunk_size = 1000
            total_records = len(product_data)
            chunks = []
            
            for i in range(0, total_records, chunk_size):
                chunk_end = min(i + chunk_size, total_records)
                chunk_data = product_data.iloc[i:chunk_end]
                chunks.append(chunk_data)

            print(f"  Split {total_records} records into {len(chunks)} chunks")

            # Process each chunk
            for chunk_idx, chunk_data in enumerate(chunks, 1):
                chunk_filename = self.stock_output.replace('.xlsx', f'_{chunk_idx}.xlsx')
                print(f"  Creating chunk {chunk_idx}/{len(chunks)}: {len(chunk_data)} records -> {chunk_filename}")

                # Copy template
                shutil.copy2(self.stock_template, chunk_filename)

                # Load workbook
                workbook = load_workbook(chunk_filename)
                sheet = workbook.active

                # Prepare data for this chunk
                stock_data = []
                for sku in chunk_data['SKU']:
                    if pd.notna(sku) and str(sku).strip() != '':
                        # Get stock from PRICES.XLS
                        stock_qty = self.get_stock_for_sku(str(sku))

                        stock_data.append({
                            'SKU': str(sku),
                            'SKU ID': '',  # Leave blank as requested
                            'New quantity': stock_qty
                        })

                # Write data starting from row 3 (row 2 has headers)
                for i, row_data in enumerate(stock_data, 3):
                    sheet.cell(row=i, column=1, value=row_data['SKU'])
                    sheet.cell(row=i, column=2, value=row_data['SKU ID'])
                    sheet.cell(row=i, column=3, value=row_data['New quantity'])

                # Save workbook
                workbook.save(chunk_filename)
                workbook.close()

                print(f"    Created chunk {chunk_idx} with {len(stock_data)} records")

            print(f"  Created {len(chunks)} stock update files")
            return True

        except Exception as e:
            print(f"Error creating stock update files: {e}")
            return False

    def process_updates(self, product_data, base_prices):
        """Process both price and stock updates"""
        print("\nProcessing price and stock updates from PRICES.XLS...")

        success_price = self.create_price_update_file(product_data, base_prices)
        success_stock = self.create_stock_update_file(product_data)

        if success_price and success_stock:
            print("✅ Price and stock update files created successfully")
            print(f"  Price files: {self.price_output} (chunked)")
            print(f"  Stock files: {self.stock_output} (chunked)")
        else:
            print("❌ Some update files failed to create")

        return success_price and success_stock

def test_price_stock_updater():
    """Test the price and stock updater"""
    updater = PriceStockUpdater()

    # Test data
    test_skus = ['HBG104955BL', 'HBG104955G', 'TEST123']
    test_base_prices = [19.99, 24.99, 15.99]

    print("Testing price lookup:")
    for sku in test_skus:
        price = updater.get_price_for_sku(sku)
        stock = updater.get_stock_for_sku(sku)
        print(f"  {sku}: Price={price}, Stock={stock}")

    # Test file creation
    test_data = pd.DataFrame({
        'SKU': test_skus,
        'Product Name': ['Test Product 1', 'Test Product 2', 'Test Product 3']
    })

    updater.process_updates(test_data, test_base_prices)

if __name__ == "__main__":
    test_price_stock_updater() 