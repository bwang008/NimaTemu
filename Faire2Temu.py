import pandas as pd
import shutil
from openpyxl import load_workbook
import warnings
import re
import argparse
import sys
import math

# Import the PriceStockUpdater
from price_stock_updater import PriceStockUpdater

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def copy_mapped_data(filter_stock=True):
    """
    Enhanced tool to copy mapped data from Faire products to Temu template.
    
    Uses a configurable mapping dictionary to copy multiple columns.
    
    CATEGORY SPLITTING:
    This script now supports splitting products into multiple categories based on SKU prefixes.
    
    To add new categories, modify the CATEGORY_CONFIGS dictionary below or use the
    add_category_config() function before calling copy_mapped_data().
    
    Example of adding a new category:
        add_category_config('hats', ['CAP', 'HAT'], 'output/temu_template_hats.xlsx', 'Hats and Caps')
        copy_mapped_data()
    
    The 'other' category is always the catch-all for products not matching any specific prefixes.
    
    CHUNKING:
    All output files are now split into chunks of 1000 records each to comply with Temu's requirements.
    
    Args:
        filter_stock (bool): If True, only process products with stock > 0. Default is True.
    """
    
    # ============================================================================
    # COLUMN MAPPING DICTIONARY
    # ============================================================================
    # Configure your column mappings here:
    # Key: Faire column name (from faire_products.xlsx)
    # Value: Temu column name (from temu_template.xlsx)
    
    COLUMN_MAPPINGS = {
        # Basic product information
        'Product Name (English)': 'Product Name',
        'Description (English)': 'Product Description',
        'SKU': 'Contribution SKU',
        # Note: USD Unit Retail Price is now used for pricing strategy calculation
        # 'USD Unit Retail Price': 'Base Price - USD',  # REMOVED - handled by pricing strategy
        'On Hand Inventory': 'Quantity',
        'Made In Country': 'Country/Region of Origin',
        
        # Note: Contribution Goods will be handled separately with SKU transformation
        
        # Option mappings
        'Option 1 Name': 'Variation Theme',
        'Option 1 Value': 'Color',
        
        # Price mappings
        # Note: USD Unit Retail Price is now used for pricing strategy calculation
        # 'USD Unit Retail Price': 'List Price - USD',  # REMOVED - handled by pricing strategy
        
        # Dimension mappings
        'Item Weight': 'Weight - lb',
        'Item Length': 'Length - in',
        'Item Width': 'Width - in',
        'Item Height': 'Height - in',
        
        # Optional mappings - uncomment and modify as needed
        # 'Product Status': 'Status',
        # 'Product Type': 'Category',
        # 'Product Images': 'Detail Images URL',
        # 'Option 2 Name': 'Size',
        # 'Option 2 Value': 'Size Value',
    }
    
    # ============================================================================
    # FIXED COLUMN VALUES DICTIONARY
    # ============================================================================
    # Configure fixed values for specific Temu columns here:
    # Key: Temu column name (from temu_template.xlsx)
    # Value: Fixed value to assign to all rows
    
    FIXED_COLUMN_VALUES = {
        'Category': '29153',
        'Country/Region of Origin': 'Mainland China',
        'Province of Origin': 'Guangdong',
        'Update or Add': 'Add',
        'Shipping Template': 'NIMA2',
        # Note: Size is now handled by conditional logic
        # 'Size': 'One Size',  # REMOVED - handled by conditional logic
        'California Proposition 65 Warning Type': 'No Warning Applicable',
        
        # Add more fixed values as needed:
        # 'Status': 'Active',
        # 'Brand': 'Your Brand Name',
        # 'Handling Time': '1',
        # 'Import Designation': 'General',
        # 'Fulfillment Channel': 'FBA',
    }
    
    # ============================================================================
    # DATA TRANSFORMATION FUNCTIONS
    # ============================================================================
    
    def transform_price(price_value):
        """Transform price values to ensure they are numeric"""
        if pd.isna(price_value) or price_value == '':
            return ''
        try:
            return float(price_value)
        except (ValueError, TypeError):
            return ''
    
    def transform_product_name(name):
        """Transform product names to ensure they are strings"""
        if pd.isna(name) or name == '':
            return ''
        return str(name).strip()
    
    def transform_sku_to_goods(sku_value):
        """Transform SKU to Contribution Goods by removing non-numeric suffixes"""
        if pd.isna(sku_value) or sku_value == '':
            return ''
        
        sku_str = str(sku_value).strip()
        
        # Remove common non-numeric suffixes (2-3 letters at the end)
        # This handles cases like 'HBG104955BL' -> 'HBG104955'
        if len(sku_str) > 3:
            # Check if the last 2-3 characters are letters
            suffix = sku_str[-3:] if len(sku_str) >= 3 else sku_str[-2:]
            if suffix.isalpha() and len(suffix) >= 2:
                # Remove the suffix
                base_sku = sku_str[:-len(suffix)]
                return base_sku
        
        return sku_str
    
    def split_image_urls(image_urls_str):
        """Split image URLs and return the first one"""
        if pd.isna(image_urls_str) or image_urls_str == '':
            return ''
        
        # Convert to string and split by common delimiters
        urls_str = str(image_urls_str)
        
        # Split by common delimiters (comma, semicolon, pipe, newline)
        delimiters = [',', ';', '|', '\n', '\r\n']
        for delimiter in delimiters:
            if delimiter in urls_str:
                urls = urls_str.split(delimiter)
                # Return the first non-empty URL
                for url in urls:
                    url_clean = url.strip()
                    if url_clean and url_clean != '':
                        return url_clean
                return ''
        
        # If no delimiters found, return the whole string
        return urls_str.strip()
    
    # ============================================================================
    # TRANSFORMATIONS DICTIONARY
    # ============================================================================
    # Map column names to their transformation functions
    
    TRANSFORMATIONS = {
        'Product Name (English)': transform_product_name,
        'Description (English)': transform_product_name,
        'USD Unit Retail Price': transform_price,
    }
    
    # ============================================================================
    # CATEGORY ASSIGNMENT SYSTEM
    # ============================================================================
    
    def add_category_config(category_name, prefixes, output_file, description):
        """Add a new category configuration"""
        global CATEGORY_CONFIGS
        if 'CATEGORY_CONFIGS' not in globals():
            CATEGORY_CONFIGS = {}
        CATEGORY_CONFIGS[category_name] = {
            'prefixes': prefixes,
            'output_file': output_file,
            'description': description
        }
        print(f"Added category: {category_name} with prefixes {prefixes}")
    
    class CategoryAssigner:
        """Enhanced category assignment system"""
        
        def __init__(self):
            self.category_rules = {
                '29153': {
                    'description': 'Unknown',
                    'keywords': []
                },
                '2062': {
                    'description': 'Pet Supplies / Small Animals / Carriers',
                    'keywords': ['pet', 'dog', 'cat', 'animal', 'carrier', 'leash', 'collar']
                },
                '29312': {
                    'description': 'Clothing, Shoes & Jewelry / Women / Accessories / Sunglasses & Eyewear',
                    'keywords': ['sunglasses', 'eyewear', 'glasses', 'optical', 'vision']
                },
                '29163': {
                    'description': 'Tote bags and totes',
                    'keywords': ['tote', 'bag', 'handbag', 'purse', 'clutch', 'wallet']
                },
                '30988': {
                    'description': 'Clothing, Shoes & Jewelry / Luggage & Travel Gear / Cosmetic Cases',
                    'keywords': ['cosmetic', 'makeup', 'beauty', 'travel', 'case']
                },
                '29324': {
                    'description': 'Clothing, Shoes & Jewelry / Women / Accessories / Wallets',
                    'keywords': ['wallet', 'purse', 'clutch', 'coin', 'card']
                },
                '29264': {
                    'description': 'Clothing, Shoes & Jewelry / Women / Accessories / Belts',
                    'keywords': ['belt', 'waist', 'accessory']
                },
                '29164': {
                    'description': 'Backpacks',
                    'keywords': ['backpack', 'school', 'laptop', 'travel']
                },
                '24380': {
                    'description': 'Cell Phones & Accessories / Cases, Holsters & Sleeves',
                    'keywords': ['phone', 'mobile', 'case', 'protector', 'cover']
                },
                '29542': {
                    'description': 'Clothing, Shoes & Jewelry / Women / Jewelry / Necklaces',
                    'keywords': ['necklace', 'jewelry', 'pendant', 'chain']
                },
                '36256': {
                    'description': 'Sports & Outdoors / Sports / Leisure Sports / Pickleball / Paddles',
                    'keywords': ['pickleball', 'paddle', 'sport', 'game']
                },
                '19843': {
                    'description': 'Beauty & Personal Care / Foot, Hand & Nail Care / Tools & Accessories',
                    'keywords': ['nail', 'beauty', 'care', 'tool']
                },
                '29522': {
                    'description': 'Clothing, Shoes & Jewelry / Women / Jewelry / Brooches & Pins',
                    'keywords': ['brooch', 'pin', 'jewelry', 'accessory']
                }
            }
        
        def determine_category(self, product_name, image_data=None):
            """Determine the best category based on product name and image data"""
            if not product_name:
                return '29153'  # Default category
            
            product_name_lower = product_name.lower()
            
            # Check each category's keywords
            for category_code, rule in self.category_rules.items():
                for keyword in rule['keywords']:
                    if keyword.lower() in product_name_lower:
                        return category_code
            
            return '29153'  # Default if no match found
        
        def get_category_info(self, category_code):
            """Get category information by code"""
            return self.category_rules.get(category_code, None)
    
    # ============================================================================
    # UTILITY FUNCTIONS
    # ============================================================================
    
    def split_data_into_chunks(data, chunk_size=1000):
        """Split data into chunks of specified size"""
        chunks = []
        for i in range(0, len(data), chunk_size):
            chunks.append(data[i:i + chunk_size])
        return chunks
    
    def generate_chunk_filename(base_filename, chunk_number):
        """Generate filename for a specific chunk"""
        name, ext = base_filename.rsplit('.', 1)
        return f"{name}_{chunk_number}.{ext}"
    
    # ============================================================================
    # PROCESSING FUNCTION FOR EACH CATEGORY
    # ============================================================================
    
    def process_product_category(product_data, template_file, output_file, category_name):
        """Process a specific category of products and save to output file with chunking"""
        
        print(f"Processing {category_name} category ({len(product_data)} products)...")
        
        # Split data into chunks of 1000 records
        data_chunks = split_data_into_chunks(product_data, 1000)
        print(f"  Split into {len(data_chunks)} chunks of max 1000 records each")
        
        # Process each chunk
        for chunk_idx, chunk_data in enumerate(data_chunks, 1):
            chunk_filename = generate_chunk_filename(output_file, chunk_idx)
            print(f"  Processing chunk {chunk_idx}/{len(data_chunks)}: {len(chunk_data)} products -> {chunk_filename}")
            
            # Copy the template file to chunk output
            shutil.copy2(template_file, chunk_filename)
            
            # Load the copied workbook and modify it
            workbook = load_workbook(chunk_filename)
            template_sheet = workbook['Template']
            
            # Convert chunk_data back to DataFrame for easier processing
            category_df = pd.DataFrame(chunk_data)
            
            # Find all columns named 'Quantity' in the template
            quantity_col_indices = []
            for col_idx, cell in enumerate(template_sheet[2], 1):
                if str(cell.value) == 'Quantity':
                    quantity_col_indices.append(col_idx)
            
            # Find all columns named 'Base Price - USD' and 'List Price - USD'
            base_price_col_indices = []
            list_price_col_indices = []
            for col_idx, cell in enumerate(template_sheet[2], 1):
                if str(cell.value) == 'Base Price - USD':
                    base_price_col_indices.append(col_idx)
                if str(cell.value) == 'List Price - USD':
                    list_price_col_indices.append(col_idx)
            
            # Find Variation Theme and Size columns for conditional logic
            variation_theme_col_idx = None
            size_col_idx = None
            for col_idx, cell in enumerate(template_sheet[2], 1):
                if str(cell.value) == 'Variation Theme':
                    variation_theme_col_idx = col_idx
                if str(cell.value) == 'Size':
                    size_col_idx = col_idx
            
            # Process each mapping
            for faire_col, temu_col in COLUMN_MAPPINGS.items():
                if faire_col in category_df.columns:
                    print(f"    Mapping: {faire_col} -> {temu_col}")
                    
                    # Get source data
                    source_data = category_df[faire_col].tolist()
                    
                    # Find the column index in Temu template
                    temu_col_idx = None
                    for col_idx, cell in enumerate(template_sheet[2], 1):  # Row 2 contains headers
                        if temu_col in str(cell.value):
                            temu_col_idx = col_idx
                            break
                    
                    if temu_col_idx is None:
                        print(f"      Warning: Could not find column '{temu_col}' in template")
                        continue
                    
                    # Apply transformation if defined
                    if faire_col in TRANSFORMATIONS:
                        source_data = [TRANSFORMATIONS[faire_col](value) for value in source_data]
                        print(f"      Applied transformation: {faire_col}")
                    
                    # Special handling for Quantity: populate all Quantity columns
                    if temu_col == 'Quantity' and quantity_col_indices:
                        for row_idx, value in enumerate(source_data, 5):
                            # Handle NaN values
                            cell_value = '' if pd.isna(value) else value
                            for q_col_idx in quantity_col_indices:
                                template_sheet.cell(row=row_idx, column=q_col_idx, value=cell_value)
                        print(f"      Copied {len(source_data)} values to all Quantity columns ({len(quantity_col_indices)})")
                        continue  # Skip the default single-column write below
                    
                    # Write data to template (default: single column)
                    for row_idx, value in enumerate(source_data, 5):
                        # Handle NaN values
                        if pd.isna(value):
                            template_sheet.cell(row=row_idx, column=temu_col_idx, value='')
                        else:
                            template_sheet.cell(row=row_idx, column=temu_col_idx, value=value)
                    
                    print(f"      Copied {len(source_data)} values")
                else:
                    print(f"    Skipping: {faire_col} -> {temu_col} (column not found)")
            
            # Conditional logic for Variation Theme and Color assignment
            print("    Processing conditional Variation Theme logic...")
            if variation_theme_col_idx is not None:
                # Get the Contribution Goods data to detect duplicates
                contribution_goods_data = []
                if 'SKU' in category_df.columns:
                    # Transform SKU to Contribution Goods for comparison
                    sku_data = category_df['SKU'].tolist()
                    contribution_goods_data = [transform_sku_to_goods(sku) for sku in sku_data]
                
                # Find the Color column
                color_col_idx = None
                for col_idx, cell in enumerate(template_sheet[2], 1):
                    if str(cell.value) == 'Color':
                        color_col_idx = col_idx
                        break
                
                # Get the Color data that was mapped
                color_data = []
                if 'Option 1 Value' in category_df.columns:
                    color_data = category_df['Option 1 Value'].tolist()
                
                if color_col_idx is not None and contribution_goods_data and color_data:
                    # Count occurrences of each Contribution Goods value
                    goods_count = {}
                    for goods in contribution_goods_data:
                        goods_count[goods] = goods_count.get(goods, 0) + 1
                    
                    # Process each row for conditional logic
                    goods_occurrence = {}  # Track occurrence count for each goods value
                    
                    for row_idx, (color_value, goods_value) in enumerate(zip(color_data, contribution_goods_data), 5):
                        if pd.isna(color_value) or str(color_value).strip() == '':
                            # No color value - determine if this is part of a multi-variant product
                            if goods_count.get(goods_value, 0) > 1:
                                # Multiple variants exist - assign sequential color
                                goods_occurrence[goods_value] = goods_occurrence.get(goods_value, 0) + 1
                                color_number = goods_occurrence[goods_value]
                                template_sheet.cell(row=row_idx, column=variation_theme_col_idx, value='Color')
                                template_sheet.cell(row=row_idx, column=color_col_idx, value=f'Color {color_number}')
                            else:
                                # Single variant - use 'One Color'
                                template_sheet.cell(row=row_idx, column=variation_theme_col_idx, value='Color')
                                template_sheet.cell(row=row_idx, column=color_col_idx, value='One Color')
                        else:
                            # Has color value - Variation Theme stays as 'Color' (already set by mapping)
                            # Color value is already set by the mapping
                            pass
                    
                    print(f"      Applied conditional logic to {len(color_data)} rows")
                    print(f"      - Rows with color: Set Variation Theme = 'Color' (existing value)")
                    print(f"      - Rows without color: Set Variation Theme = 'Color'")
                    print(f"        - Multi-variant products: Sequential 'Color 1', 'Color 2', etc.")
                    print(f"        - Single products: 'One Color'")
            
            # Pricing strategy calculation
            print("    Calculating pricing strategy (1x and 1.25x Faire price, floored to X.99)...")
            if base_price_col_indices and list_price_col_indices:
                # Get the USD Unit Retail Price data
                if 'USD Unit Retail Price' in category_df.columns:
                    price_data = category_df['USD Unit Retail Price'].tolist()
                    
                    for row_idx, price_value in enumerate(price_data, 5):
                        if pd.notna(price_value) and price_value != '':
                            try:
                                # Convert to float and calculate pricing strategy
                                price_float = float(price_value)
                                
                                # Base Price: 1x Faire price, floored to X.99
                                base_price = math.floor(price_float) - 0.01
                                base_price = max(0.01, base_price)  # Ensure minimum price
                                
                                # List Price: 1.25x Faire price, floored to X.99
                                list_price = math.floor(price_float * 1.25) - 0.01
                                list_price = max(base_price + 0.01, list_price)  # Ensure list price > base price
                                
                                # Write to all Base Price columns
                                for col_idx in base_price_col_indices:
                                    template_sheet.cell(row=row_idx, column=col_idx, value=base_price)
                                
                                # Write to all List Price columns
                                for col_idx in list_price_col_indices:
                                    template_sheet.cell(row=row_idx, column=col_idx, value=list_price)
                                    
                            except (ValueError, TypeError):
                                # If price conversion fails, skip this row
                                continue
                    
                    print(f"      Set pricing strategy for {len(price_data)} rows")
                    print(f"      Base Price: 1x Faire price, floored, minus 1 cent")
                    print(f"      List Price: 1.25x Faire price, floored, minus 1 cent")
            
            # Contribution Goods transformation
            print("    Processing Contribution Goods transformation...")
            if 'SKU' in category_df.columns:
                sku_data = category_df['SKU'].tolist()
                contribution_goods_data = [transform_sku_to_goods(sku) for sku in sku_data]
                
                # Find the Contribution Goods column
                contribution_goods_col_idx = None
                for col_idx, cell in enumerate(template_sheet[2], 1):
                    if str(cell.value) == 'Contribution Goods':
                        contribution_goods_col_idx = col_idx
                        break
                
                if contribution_goods_col_idx is not None:
                    for row_idx, goods_value in enumerate(contribution_goods_data, 5):
                        template_sheet.cell(row=row_idx, column=contribution_goods_col_idx, value=goods_value)
                    
                    print(f"      Transformed {len(contribution_goods_data)} SKUs to Contribution Goods")
                    print("      Sample transformations:")
                    for i, (sku, goods) in enumerate(zip(sku_data[:5], contribution_goods_data[:5]), 1):
                        print(f"        {i}. '{sku}' → '{goods}'")
            
            # Image URL processing
            print("    Processing Image URLs...")
            image_columns = [col for col in category_df.columns if 'Image' in col]
            if image_columns:
                print(f"      Found {len(image_columns)} image columns")
                
                # Find Detail Images URL columns in template
                detail_images_col_indices = []
                for col_idx, cell in enumerate(template_sheet[2], 1):
                    if 'Detail Images URL' in str(cell.value):
                        detail_images_col_indices.append(col_idx)
                
                if detail_images_col_indices:
                    print(f"      Found {len(detail_images_col_indices)} Detail Images URL columns")
                    
                    option_image_count = 0
                    product_image_count = 0
                    no_image_count = 0
                    
                    for row_idx, row in enumerate(category_df.iterrows(), 5):
                        row_data = row[1]
                        
                        # Try to find image data
                        image_urls = None
                        
                        # First, try Option Image columns
                        for col in image_columns:
                            if 'Option' in col and pd.notna(row_data[col]) and str(row_data[col]).strip() != '':
                                image_urls = row_data[col]
                                option_image_count += 1
                                break
                        
                        # If no option image, try Product Images
                        if image_urls is None:
                            for col in image_columns:
                                if 'Product' in col and pd.notna(row_data[col]) and str(row_data[col]).strip() != '':
                                    image_urls = row_data[col]
                                    product_image_count += 1
                                    break
                        
                        if image_urls is None:
                            no_image_count += 1
                            continue
                        
                        # Process image URLs
                        processed_urls = split_image_urls(image_urls)
                        
                        # Write to all Detail Images URL columns
                        for col_idx in detail_images_col_indices:
                            template_sheet.cell(row=row_idx, column=col_idx, value=processed_urls)
                    
                    print(f"        - Used Option Image: {option_image_count} rows")
                    print(f"        - Used Product Images: {product_image_count} rows")
                    print(f"        - No image data: {no_image_count} rows")
            else:
                print("      Warning: No image columns found in Faire file")
            
            # Process fixed column values with conditional category assignment
            print("    Processing fixed column values with conditional category assignment...")
            
            for temu_col, fixed_value in FIXED_COLUMN_VALUES.items():
                print(f"      Setting fixed value: {temu_col} = '{fixed_value}'")
                
                # Find the column index in Temu template
                temu_col_idx = None
                for col_idx, cell in enumerate(template_sheet[2], 1):  # Row 2 contains headers
                    if temu_col in str(cell.value):
                        temu_col_idx = col_idx
                        break
                
                if temu_col_idx is None:
                    print(f"        Warning: Could not find column '{temu_col}' in template")
                    continue
                
                # Get the number of data rows
                num_data_rows = len(category_df)
                
                # Special handling for Category column with enhanced assignment
                if temu_col == 'Category':
                    print("        Applying enhanced category assignment...")
                    category_assignments = {}
                    
                    # Get product names and image data for category assignment
                    product_names = []
                    image_data = []
                    if 'Product Name (English)' in category_df.columns:
                        product_names = category_df['Product Name (English)'].tolist()
                    if 'Product Images' in category_df.columns:
                        image_data = category_df['Product Images'].tolist()
                    
                    # Process each row with enhanced category assignment
                    for row_idx in range(5, 5 + num_data_rows):
                        product_name = product_names[row_idx - 5] if row_idx - 5 < len(product_names) else ''
                        img_data = image_data[row_idx - 5] if row_idx - 5 < len(image_data) else None
                        
                        # Use the enhanced category assigner
                        category_code = category_assigner.determine_category(product_name, img_data)
                        
                        # Track category assignments for reporting
                        if category_code not in category_assignments:
                            category_assignments[category_code] = 0
                        category_assignments[category_code] += 1
                        
                        template_sheet.cell(row=row_idx, column=temu_col_idx, value=category_code)
                    
                    # Report category assignments with descriptions
                    print(f"        Category assignments:")
                    for category_code, count in category_assignments.items():
                        category_info = category_assigner.get_category_info(category_code)
                        description = category_info['description'] if category_info else 'Unknown'
                        print(f"          {category_code} ({description}): {count} products")
                else:
                    # Write fixed value to all data rows for non-category columns
                    for row_idx in range(5, 5 + num_data_rows):
                        template_sheet.cell(row=row_idx, column=temu_col_idx, value=fixed_value)
                
                print(f"        Set values for {num_data_rows} rows")
            
            # Save the workbook
            workbook.save(chunk_filename)
            workbook.close()
            
            print(f"      Completed chunk {chunk_idx}/{len(data_chunks)}: {chunk_filename}")
        
        print(f"Completed processing {category_name} category ({len(product_data)} products) -> {len(data_chunks)} files")
    
    # ============================================================================
    # MAIN PROCESSING FUNCTION
    # ============================================================================
    
    try:
        # File paths
        faire_file = 'data/faire_products.xlsx'
        temu_template_file = 'data/temu_template.xlsx'
        
        # Define category configurations
        CATEGORY_CONFIGS = {
            'handbags': {
                'prefixes': ['HBG', 'HW', 'HM', 'HL'],
                'output_file': 'output/temu_template_handbags.xlsx',
                'description': 'Handbags, Wallets, Cosmetic Bags, Travel Bags'
            },
            'other': {
                'prefixes': [],  # Empty means catch-all for anything not in other categories
                'output_file': 'output/temu_template_other.xlsx',
                'description': 'All other products (hats, accessories, etc.)'
            }
            # Future categories can be added here:
            # 'hats': {
            #     'prefixes': ['CAP', 'HAT'],
            #     'output_file': 'output/temu_template_hats.xlsx',
            #     'description': 'Hats and Caps'
            # },
            # 'accessories': {
            #     'prefixes': ['TO-', 'ACC'],
            #     'output_file': 'output/temu_template_accessories.xlsx',
            #     'description': 'Accessories and Straps'
            # }
        }
        
        print("Starting enhanced mapping tool with chunking...")
        print(f"Mapping {len(COLUMN_MAPPINGS)} columns")
        print(f"Setting {len(FIXED_COLUMN_VALUES)} fixed values")
        print(f"Categories: {list(CATEGORY_CONFIGS.keys())}")
        print("Chunking: All files will be split into chunks of 1000 records")
        
        # Initialize category assigner
        category_assigner = CategoryAssigner()
        print(f"Enhanced category rules: {len(category_assigner.category_rules)} categories available")
        
        # Step 1: Load Faire products file
        print("Loading Faire products file...")
        faire_df = pd.read_excel(faire_file, sheet_name='Products')
        
        # Step 2: Load Temu template file
        print("Loading Temu template...")
        temu_df = pd.read_excel(temu_template_file, sheet_name='Template', header=1)
        
        # Step 3: Validate mappings
        print("Validating column mappings...")
        missing_faire_columns = []
        missing_temu_columns = []
        
        for faire_col, temu_col in COLUMN_MAPPINGS.items():
            if faire_col not in faire_df.columns:
                missing_faire_columns.append(faire_col)
            if temu_col not in temu_df.columns:
                missing_temu_columns.append(temu_col)
        
        if missing_faire_columns:
            print(f"Warning: Missing Faire columns: {missing_faire_columns}")
        if missing_temu_columns:
            print(f"Warning: Missing Temu columns: {missing_temu_columns}")
        
        # Step 4: Filter and split data into categories
        print("Filtering and splitting data into categories...")
        
        # Get data from row 4 onwards (skip header rows)
        data_df = faire_df.iloc[3:].copy()
        
        # Filter for products with stock > 0 (if enabled)
        if filter_stock:
            print("Filtering products with stock > 0...")
            total_products = len(data_df)
            
            # Filter based on 'On Hand Inventory' > 0
            if 'On Hand Inventory' in data_df.columns:
                # Convert to numeric, handling any non-numeric values
                inventory_data = pd.to_numeric(data_df['On Hand Inventory'], errors='coerce')
                in_stock_mask = inventory_data > 0
                data_df = data_df[in_stock_mask]
                
                filtered_products = len(data_df)
                print(f"  Total products: {total_products}")
                print(f"  Products with stock > 0: {filtered_products}")
                print(f"  Products filtered out: {total_products - filtered_products}")
            else:
                print("  Warning: 'On Hand Inventory' column not found, processing all products")
        else:
            print("Stock filtering disabled - processing all products")
        
        # Initialize category data containers
        category_data = {category: [] for category in CATEGORY_CONFIGS.keys()}
        
        # Split data based on SKU prefixes
        for idx, row in data_df.iterrows():
            sku = str(row['SKU']) if pd.notna(row['SKU']) else ''
            assigned_category = None
            
            # Check each category's prefixes (except 'other' which is catch-all)
            for category, config in CATEGORY_CONFIGS.items():
                if category == 'other':
                    continue  # Skip 'other' for now, it's the catch-all
                
                if any(sku.startswith(prefix) for prefix in config['prefixes']):
                    assigned_category = category
                    break
            
            # If no specific category found, assign to 'other'
            if assigned_category is None:
                assigned_category = 'other'
            
            category_data[assigned_category].append(row)
        
        # Print category breakdown
        print("Category breakdown:")
        for category, data in category_data.items():
            config = CATEGORY_CONFIGS[category]
            print(f"  {category.title()}: {len(data)} products ({config['description']})")
        
        # Process each category
        for category, data in category_data.items():
            if len(data) > 0:  # Only process categories with data
                config = CATEGORY_CONFIGS[category]
                print(f"\nProcessing {category} products...")
                process_product_category(data, temu_template_file, config['output_file'], category)
        
        print(f"\nSuccess! Output files saved to:")
        for category, config in CATEGORY_CONFIGS.items():
            if len(category_data[category]) > 0:
                print(f"  {category.title()}: {config['output_file']} (chunked)")
        
        # ============================================================================
        # PRICE AND STOCK UPDATE PROCESSING
        # ============================================================================
        print("\n" + "="*60)
        print("PROCESSING PRICE AND STOCK UPDATES FROM PRICES.XLS")
        print("="*60)
        
        try:
            # Initialize the price/stock updater
            updater = PriceStockUpdater()
            
            # Prepare product data for price/stock updates
            # We need all SKUs and their base prices from the processed data
            all_skus = []
            all_base_prices = []
            
            # Collect all SKUs and base prices from all categories
            for category, data in category_data.items():
                if len(data) > 0:
                    for row in data:
                        sku = str(row['SKU']) if pd.notna(row['SKU']) else ''
                        base_price = row.get('USD Unit Retail Price', 0)
                        
                        if sku and sku.strip() != '':
                            all_skus.append(sku)
                            all_base_prices.append(base_price)
            
            # Create product data DataFrame for the updater
            product_data = pd.DataFrame({
                'SKU': all_skus,
                'Product Name': ['Product'] * len(all_skus)  # Placeholder, not used by updater
            })
            
            print(f"Processing price and stock updates for {len(all_skus)} products...")
            
            # Process the updates
            success = updater.process_updates(product_data, all_base_prices)
            
            if success:
                print("✅ Price and stock update files created successfully!")
                print(f"  Price files: {updater.price_output} (chunked)")
                print(f"  Stock files: {updater.stock_output} (chunked)")
            else:
                print("❌ Some update files failed to create")
                
        except Exception as e:
            print(f"❌ Error during price/stock update processing: {e}")
            print("Continuing with main process...")
        
        return  # Exit early since we're now using separate function
        
    except FileNotFoundError as e:
        print(f"Error: Could not find a file. Please check your file paths. Details: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")

def show_available_columns():
    """Show available columns in both files for reference."""
    try:
        faire_df = pd.read_excel('data/faire_products.xlsx', sheet_name='Products')
        temu_df = pd.read_excel('data/temu_template.xlsx', sheet_name='Template', header=1)
        
        print("AVAILABLE COLUMNS FOR MAPPING:")
        print("=" * 60)
        print("FAIRE COLUMNS:")
        for i, col in enumerate(faire_df.columns, 1):
            print(f"  {i:2d}. {col}")
        
        print("\nTEMU COLUMNS:")
        for i, col in enumerate(temu_df.columns, 1):
            print(f"  {i:2d}. {col}")
            
    except Exception as e:
        print(f"Error showing columns: {e}")

def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(
        description='Faire to Temu data migration tool with category splitting, stock filtering, and chunking',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python Faire2Temu.py                    # Default: filter stock > 0, chunk files
  python Faire2Temu.py --filter-stock     # Explicitly enable stock filtering
  python Faire2Temu.py --no-filter-stock  # Disable stock filtering (process all products)
  python Faire2Temu.py -f                 # Short form: enable stock filtering
  python Faire2Temu.py -F                 # Short form: disable stock filtering
        """
    )
    
    parser.add_argument(
        '--filter-stock', 
        action='store_true',
        default=True,
        help='Filter products with stock > 0 (default: True)'
    )
    parser.add_argument(
        '--no-filter-stock', 
        action='store_true',
        help='Disable stock filtering (process all products)'
    )
    parser.add_argument(
        '-f', '--force-filter',
        action='store_true',
        help='Force enable stock filtering'
    )
    parser.add_argument(
        '-F', '--force-no-filter',
        action='store_true',
        help='Force disable stock filtering'
    )
    
    return parser.parse_args()

if __name__ == "__main__":
    args = parse_arguments()
    
    # Determine filtering behavior
    if args.no_filter_stock or args.force_no_filter:
        filter_stock = False
    elif args.filter_stock or args.force_filter:
        filter_stock = True
    else:
        filter_stock = True  # Default behavior
    
    print(f"Stock filtering: {'ENABLED' if filter_stock else 'DISABLED'}")
    copy_mapped_data(filter_stock=filter_stock) 