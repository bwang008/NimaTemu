import pandas as pd
import numpy as np
import shutil
from openpyxl import load_workbook
import warnings

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

print("Running script...  ")

def migrate_faire_to_temu(faire_path, temu_template_path, output_path):
    """
    Reads product data from a Faire export file, transforms it, and maps it
    to a Temu upload template.

    Args:
        faire_path (str): The file path for the Faire product export (Excel).
        temu_template_path (str): The file path for the Temu template (Excel).
        output_path (str): The path to save the generated Temu upload file (Excel).
    """
    
    try:
        # --- 1. Load the Source and Template Files ---
        # Load the Faire product data from the "Products" sheet. We skip the second row which contains non-header info.
        faire_df = pd.read_excel(faire_path, sheet_name="Products", header=0, skiprows=[1])
        print(f"Successfully loaded {len(faire_df)} rows from Faire file.")

        # Load the Temu template from the "Template" sheet. The actual headers are in the second row (index 1).
        temu_template_df = pd.read_excel(temu_template_path, sheet_name="Template", header=1)
        print("Temu template loaded successfully.")
        
        # --- 1a. Load Column Mappings ---
        # Read the Faire column names from row 1 (header=0)
        faire_columns_df = pd.read_excel(faire_path, sheet_name="Products", header=0, nrows=1)
        faire_columns = faire_columns_df.columns.tolist()
        
        # Read the Temu column names from row 2 (header=1)
        temu_columns_df = pd.read_excel(temu_template_path, sheet_name="Template", header=1, nrows=1)
        temu_columns = temu_columns_df.columns.tolist()
        
        # Create mapping dictionary based on the actual column names
        # Map Faire column names to Temu column names
        column_mapping = {
            'Product Name': 'Product Name (English)',  # Faire product name to Temu Product Name
            'Product Description': 'Description (English)',  # Faire description to Temu Product Description
            'Contribution Goods': 'Product Token',  # Faire Product Token to Temu Contribution Goods
            'Contribution SKU': 'SKU',  # Faire SKU to Temu Contribution SKU
            'Base Price - USD': 'USD Unit Retail Price',  # Faire retail price to Temu Base Price
            'Quantity': 'On Hand Inventory',  # Faire inventory to Temu Quantity
            'Country/Region of Origin': 'Made In Country',  # Faire country to Temu Country/Region
            'Option 1 Name': 'Option 1 Name',  # Faire option names
            'Option 1 Value': 'Option 1 Value',  # Faire option values
            'Option 2 Name': 'Option 2 Name',
            'Option 2 Value': 'Option 2 Value',
            'Product Images': 'Product Images',  # Faire image URLs
        }

    except FileNotFoundError as e:
        print(f"Error: Could not find a file. Please check your file paths. Details: {e}")
        return
    except Exception as e:
        print(f"An error occurred during file loading. Details: {e}")
        return

    # --- 2. Process Each Product in the Faire DataFrame ---
    processed_rows = []
    for index, product in faire_df.iterrows():
        # This dictionary will hold the data for a single processed row (a single variant)
        # before it's added to our list of all processed rows.
        
        # --- 2a. Handle Products with Options (Flattening) ---
        # Check if the product has at least one option defined.
        # We check for non-null values in 'Option 1 Name' as the indicator.
        if pd.notna(product[column_mapping['Option 1 Name']]):
            # This is a product with variants. We need to create a row for each one.
            base_row = {
                # --- Map Core Product Information (Same for all variants) ---
                'Product Name': product[column_mapping['Product Name']],
                'Product Description': product[column_mapping['Product Description']],
                # Use the 'Product Token' as the parent SPU/Contribution Goods
                'Contribution Goods': product[column_mapping['Contribution Goods']], 
                'Country/Region of Origin': product[column_mapping['Country/Region of Origin']],
                # Temu requires a category ID. This is a placeholder and must be manually updated.
                'Category': '29153', # Placeholder for 'Handbags & Wallets'
            }

            # --- Split Image URLs ---
            # Split the space-separated string of image URLs into a list.
            # Fill with empty strings if no images are found to avoid errors.
            images = str(product[column_mapping['Product Images']]).split(' ') if pd.notna(product[column_mapping['Product Images']]) else [''] * 10
            # Assign images to the numerous 'Detail Images URL' columns
            for i in range(min(len(images), 30)): # Temu has many image columns
                 base_row[f'Detail Images URL.{i}' if i > 0 else 'Detail Images URL'] = images[i]

            # --- Create a Row for Each Variant ---
            # We assume options are in the same row, which is standard for Faire exports.
            # This logic would be expanded for files where variants are on separate rows.
            variant_row = base_row.copy()
            
            # --- Map Variant-Specific Information ---
            variant_row['Contribution SKU'] = product[column_mapping['Contribution SKU']]
            variant_row['Base Price - USD'] = str(product[column_mapping['Base Price - USD']]).replace('$', '')
            variant_row['Quantity'] = product[column_mapping['Quantity']]
            
            # Map Option Name and Value to Temu's 'Sale Property' columns
            variant_row['Sale Property 1'] = f"{product[column_mapping['Option 1 Name']]}:{product[column_mapping['Option 1 Value']]}"
            if pd.notna(product[column_mapping['Option 2 Name']]):
                variant_row['Sale Property 2'] = f"{product[column_mapping['Option 2 Name']]}:{product[column_mapping['Option 2 Value']]}"

            processed_rows.append(variant_row)

        else:
            # --- 2b. Handle Simple Products (No Options) ---
            # This product has no variants, so we create a single row for it.
            row_data = {
                'Product Name': product[column_mapping['Product Name']],
                'Product Description': product[column_mapping['Product Description']],
                # For simple products, the main SKU can serve as both parent and child SKU.
                'Contribution Goods': product[column_mapping['Contribution SKU']], 
                'Contribution SKU': product[column_mapping['Contribution SKU']],
                'Base Price - USD': str(product[column_mapping['Base Price - USD']]).replace('$', ''),
                'Quantity': product[column_mapping['Quantity']],
                'Country/Region of Origin': product[column_mapping['Country/Region of Origin']],
                'Category': '29153', # Placeholder
            }
            
            images = str(product[column_mapping['Product Images']]).split(' ') if pd.notna(product[column_mapping['Product Images']]) else [''] * 10
            for i in range(min(len(images), 30)):
                 row_data[f'Detail Images URL.{i}' if i > 0 else 'Detail Images URL'] = images[i]
            
            processed_rows.append(row_data)

    # --- 3. Create Final DataFrame and Save ---
    if processed_rows:
        # Concatenate all the processed row dictionaries into a single DataFrame
        temu_output_df = pd.DataFrame(processed_rows)
        
        # Reorder columns to match the original Temu template for consistency
        # and drop any columns that weren't in the original template.
        final_columns = [col for col in temu_template_df.columns if col in temu_output_df.columns]
        temu_output_df = temu_output_df[final_columns]

        # --- 4. Copy the original template and fill in data ---
        # First, copy the original template file to preserve all formatting and sheets
        # Use a temporary filename to avoid conflicts with open files
        temp_output_path = output_path.replace('.xlsx', '_temp.xlsx')
        shutil.copy2(temu_template_path, temp_output_path)
        print(f"Copied template file to: {temp_output_path}")
        
        # Load the copied workbook to modify the Template sheet
        workbook = load_workbook(temp_output_path)
        template_sheet = workbook['Template']
        
        # Clear existing data in the Template sheet (keep headers)
        # Find the last row with data
        max_row = template_sheet.max_row
        # Clear data starting from row 3 (after headers)
        if max_row > 2:
            template_sheet.delete_rows(3, max_row - 2)
        
        # Write the processed data to the Template sheet
        # Start from row 3 (after the 2 header rows)
        for row_idx, row_data in enumerate(temu_output_df.values, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                template_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Save the workbook
        workbook.save(temp_output_path)
        workbook.close()
        
        # Try to rename the temp file to the final output path
        try:
            import os
            if os.path.exists(output_path):
                os.remove(output_path)
            os.rename(temp_output_path, output_path)
        except Exception as e:
            print(f"Warning: Could not rename to final path. Using temporary file: {temp_output_path}")
            output_path = temp_output_path
        
        print(f"\nMigration complete! 🎉")
        print(f"Processed {len(temu_output_df)} total variants.")
        print(f"Output saved to: {output_path}")
        
        # --- 5. Sample Data Verification ---
        print(f"\nSample data verification:")
        print(f"First 3 rows of processed data:")
        for i, row in enumerate(temu_output_df.head(3).iterrows()):
            print(f"Row {i+1}:")
            print(f"  Product Name: {row[1].get('Product Name', 'N/A')}")
            print(f"  Contribution SKU: {row[1].get('Contribution SKU', 'N/A')}")
            print(f"  Base Price: {row[1].get('Base Price - USD', 'N/A')}")
            print(f"  Quantity: {row[1].get('Quantity', 'N/A')}")
            print()
    else:
        print("\nNo products were processed. Please check the format of your Faire file.")


# --- RUN THE SCRIPT ---
# Define the file paths for your input and output files.
# PLEASE REPLACE THESE WITH THE ACTUAL PATHS TO YOUR FILES.
faire_file = 'data/faire_products.xlsx'
temu_template_file = 'data/temu_template.xlsx'
output_file = 'output/temu_upload_generated.xlsx'

# Execute the migration function
migrate_faire_to_temu(faire_file, temu_template_file, output_file)
